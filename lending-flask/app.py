from flask import Flask, jsonify, request, render_template, session
from flask_cors import CORS
import mysql.connector
import os
from dotenv import load_dotenv
from decimal import Decimal
import json
import tempfile
import datetime
from functools import wraps
import io, calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file


load_dotenv()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, 'templates'),
            static_folder=os.path.join(BASE_DIR, 'static'))

app.secret_key = os.getenv("FLASK_SECRET_KEY", "change-this-to-something-random-and-long")
CORS(app, supports_credentials=True)

# ─── DB connection ────────────────────────────────────────────────────────────
def get_db():
    ssl_ca = os.getenv("DB_SSL_CA", "")

    ssl_ca_path = None
    if ssl_ca.strip().startswith("-----BEGIN"):
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pem", mode='w')
        tmp.write(ssl_ca)
        tmp.close()
        ssl_ca_path = tmp.name

    config = {
        "host": os.getenv("DB_HOST", "localhost"),
        "port": int(os.getenv("DB_PORT", 3306)),
        "user": os.getenv("DB_USER", "root"),
        "password": os.getenv("DB_PASSWORD", ""),
        "database": os.getenv("DB_NAME", "lending_db"),
    }
    if ssl_ca_path:
        config["ssl_ca"] = ssl_ca_path
        config["ssl_verify_cert"] = True

    return mysql.connector.connect(**config)

def decimal_default(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    if isinstance(obj, (datetime.date, datetime.datetime)):
        return str(obj)
    if isinstance(obj, datetime.timedelta):
        return str(obj)
    if isinstance(obj, bytes):
        return obj.decode('utf-8', errors='replace')
    raise TypeError(f"Object of type {type(obj)} is not JSON serializable")

def query(sql, params=(), fetchone=False, commit=False):
    conn = get_db()
    cur = conn.cursor(dictionary=True)
    cur.execute(sql, params)
    if commit:
        conn.commit()
        last_id = cur.lastrowid
        cur.close()
        conn.close()
        return last_id
    result = cur.fetchone() if fetchone else cur.fetchall()
    cur.close()
    conn.close()
    return result

# ─── Profit rate helper ───────────────────────────────────────────────────────
def get_profit_rate(slot_id):
    slot = query("SELECT slot_name FROM funding_slots WHERE id=%s", (slot_id,), fetchone=True)
    if slot and 'raigarh' in slot['slot_name'].lower():
        return 0.22
    return 0.28

# ─── Month label helper (for Excel export titles) ────────────────────────────
_MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
def monthLabel(my):
    y, m = my.split('-')
    return f"{_MONTHS[int(m)-1]} {y}"

# ─── Auth ─────────────────────────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get("role") != "admin":
            return jsonify({"error": "Admin access required"}), 403
        return f(*args, **kwargs)
    return wrapper

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"error": "Login required"}), 401
        return f(*args, **kwargs)
    return wrapper

@app.route("/api/login", methods=["POST"])
def login():
    data = request.json or {}
    pw = data.get("password", "")
    user = query("SELECT * FROM app_users WHERE password=%s", (pw,), fetchone=True)
    if not user:
        return jsonify({"error": "Invalid password"}), 401
    session["role"] = user["role"]
    session["logged_in"] = True
    return jsonify({"role": user["role"]})

@app.route("/api/session", methods=["GET"])
def check_session():
    if session.get("logged_in"):
        return jsonify({"role": session.get("role")})
    return jsonify({"role": None}), 401

@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"message": "Logged out"})

# ─── Serve frontend ───────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

# ─── SLOTS ────────────────────────────────────────────────────────────────────
@app.route("/api/slots", methods=["GET"])
@login_required
def get_slots():
    rows = query("SELECT * FROM funding_slots ORDER BY id")
    return app.response_class(
        json.dumps(rows, default=decimal_default),
        mimetype="application/json"
    )

@app.route("/api/slots/<int:slot_id>", methods=["PUT"])
@admin_required
def update_slot(slot_id):
    data = request.json
    query(
        "UPDATE funding_slots SET slot_name=%s, kothi_amount=%s WHERE id=%s",
        (data.get("slot_name"), data.get("kothi_amount"), slot_id),
        commit=True,
    )
    row = query("SELECT * FROM funding_slots WHERE id=%s", (slot_id,), fetchone=True)
    return app.response_class(json.dumps(row, default=decimal_default), mimetype="application/json")

@app.route("/api/slots/summary/<int:slot_id>/<month_year>", methods=["GET"])
@login_required
def get_summary(slot_id, month_year):
    profit_rate = get_profit_rate(slot_id)
    staff_rate  = 0.08
    net_rate    = profit_rate - staff_rate
    total_rate  = 1 + profit_rate

    row = query(
        f"""SELECT
            COUNT(*) AS total_customers,
            SUM(CASE WHEN DATE_FORMAT(c.opening_date, '%Y-%m') = %s THEN c.funding ELSE 0 END) AS this_month_funding,
            SUM(c.funding * {total_rate}) AS total_funding_and_profit,
            SUM(c.payment_has_been_done) AS actual_recovery,
            SUM(c.funding * {profit_rate}) AS total_profit_28,
            SUM(c.funding * {net_rate}) AS total_net_income,
            SUM(CASE WHEN DATE_FORMAT(c.opening_date, '%Y-%m') = %s AND fs.charges_commission = 1 THEN c.funding * {staff_rate} ELSE 0 END) AS total_staff_commission,
            SUM(c.funding) AS running_funding,
            SUM(CASE WHEN c.is_closed = 0 THEN (c.funding * {total_rate} - c.payment_has_been_done) ELSE 0 END) AS total_balance_recovery,
            SUM(CASE WHEN c.is_closed = 1 THEN c.funding ELSE 0 END) AS closed_funding_total
           FROM customers c
           JOIN funding_slots fs ON c.slot_id = fs.id
           WHERE c.slot_id=%s AND c.month_year=%s""",
        (month_year, month_year, slot_id, month_year),
        fetchone=True,
    )
    if row:
        net_profit = float(row.get("total_net_income") or 0) - float(row.get("total_staff_commission") or 0)
        row["net_profit"]           = net_profit
        row["month_year"]           = month_year
        row["slot_id"]              = slot_id
        row["profit_rate"]          = profit_rate
        row["staff_rate"]           = staff_rate
        row["net_rate"]             = net_rate
        row["closed_funding_total"] = float(row.get("closed_funding_total") or 0)
    return app.response_class(json.dumps(row, default=decimal_default), mimetype="application/json")

@app.route("/api/slots/summary", methods=["POST"])
@admin_required
def save_summary():
    data = request.json
    query(
        """INSERT INTO monthly_summary (slot_id, month_year, salary_and_other)
           VALUES (%s, %s, %s)
           ON DUPLICATE KEY UPDATE salary_and_other=%s""",
        (data["slot_id"], data["month_year"], data.get("salary_and_other", 0), data.get("salary_and_other", 0)),
        commit=True,
    )
    return jsonify({"message": "Saved"})

# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────
@app.route("/api/export/<int:slot_id>/<month_year>", methods=["GET"])
@admin_required
def export_excel(slot_id, month_year):
    slot = query("SELECT * FROM funding_slots WHERE id=%s", (slot_id,), fetchone=True)
    if not slot:
        return jsonify({"error": "Slot not found"}), 404

    profit_rate = get_profit_rate(slot_id)
    staff_rate  = 0.08
    net_rate    = profit_rate - staff_rate
    total_rate  = 1 + profit_rate

    customers = query(
        "SELECT * FROM customers WHERE slot_id=%s AND month_year=%s ORDER BY sl_no ASC",
        (slot_id, month_year),
    )
    for c in customers:
        f = float(c.get("funding") or 0)
        p = float(c.get("payment_has_been_done") or 0)
        c["profit_28_percent"]        = round(f * profit_rate, 2)
        c["staff_commission"]         = round(f * staff_rate, 2)
        c["net_income"]               = round(f * net_rate, 2)
        c["total_payment_to_be_made"] = round(f * total_rate, 2)
        c["balance_recovery"]         = round(f * total_rate - p, 2)

    summary_row = query(
        f"""SELECT
            COUNT(*) AS total_customers,
            SUM(CASE WHEN DATE_FORMAT(c.opening_date, '%Y-%m') = %s THEN c.funding ELSE 0 END) AS this_month_funding,
            SUM(c.funding * {total_rate}) AS total_funding_and_profit,
            SUM(c.payment_has_been_done) AS actual_recovery,
            SUM(c.funding * {profit_rate}) AS total_profit_28,
            SUM(c.funding * {net_rate}) AS total_net_income,
            SUM(CASE WHEN DATE_FORMAT(c.opening_date, '%Y-%m') = %s AND fs.charges_commission = 1 THEN c.funding * {staff_rate} ELSE 0 END) AS total_staff_commission,
            SUM(c.funding) AS running_funding,
            SUM(CASE WHEN c.is_closed = 0 THEN (c.funding * {total_rate} - c.payment_has_been_done) ELSE 0 END) AS total_balance_recovery,
            SUM(CASE WHEN c.is_closed = 1 THEN c.funding ELSE 0 END) AS closed_funding_total
           FROM customers c
           JOIN funding_slots fs ON c.slot_id = fs.id
           WHERE c.slot_id=%s AND c.month_year=%s""",
        (month_year, month_year, slot_id, month_year),
        fetchone=True,
    ) or {}

    collections = query(
        """SELECT *, DAY(collection_date) AS day_number, actual_collected AS amount_collected
           FROM daily_collections WHERE slot_id=%s AND month_year=%s
           ORDER BY collection_date ASC""",
        (slot_id, month_year),
    )
    opening_row = query(
        "SELECT opening_amount FROM opening_pawana WHERE slot_id=%s AND month_year=%s",
        (slot_id, month_year), fetchone=True,
    )
    opening_amount = float(opening_row["opening_amount"]) if opening_row else 0
    coll_map = {r["day_number"]: r for r in collections}
    y, m = month_year.split("-")
    days_in_month = calendar.monthrange(int(y), int(m))[1]

    daily_rows = []
    running_pawana = opening_amount
    total_daily_income = 0
    total_khata = 0
    for day in range(1, days_in_month + 1):
        rec    = coll_map.get(day)
        income = float(rec["amount_collected"]) if rec else 0
        khata  = float(rec["khata_amount"]) if rec and rec.get("khata_amount") is not None else 0
        opening_p = running_pawana
        closing_p = opening_p + income - khata
        daily_rows.append({
            "day": day, "income": income, "khata": khata,
            "opening_pawana": opening_p, "closing_pawana": closing_p,
        })
        running_pawana = closing_p
        total_daily_income += income
        total_khata += khata

    # ── Build workbook ──
    wb = Workbook()
    ws = wb.active
    ws.title = month_year

    NAVY    = "1F2D50"
    GOLD    = "C8960C"
    LIGHT   = "F4F6FB"
    WHITE_FONT   = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    TITLE_FONT   = Font(bold=True, name="Arial", size=15, color=NAVY)
    SECTION_FONT = Font(bold=True, name="Arial", size=12, color="FFFFFF")
    HEADER_FILL  = PatternFill("solid", fgColor=NAVY)
    SECTION_FILL = PatternFill("solid", fgColor=GOLD)
    TOTAL_FILL   = PatternFill("solid", fgColor="E8ECF7")
    BORDER = Border(*[Side(style="thin", color="D5D9E3")]*4)
    CENTER = Alignment(horizontal="center", vertical="center")
    RIGHT  = Alignment(horizontal="right", vertical="center")
    LEFT   = Alignment(horizontal="left", vertical="center")
    BODY_FONT = Font(name="Arial", size=10)
    BOLD_BODY = Font(name="Arial", size=10, bold=True)

    NUM_FMT = '"₹"#,##0;("₹"#,##0);"-"'
    PCT_FMT = '0.0%'

    r = 1
    ncols = 12

    def merge_title(row, text, font, fill, height=22):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER
        ws.row_dimensions[row].height = height
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = fill

    # ── Report title ──
    merge_title(r, f"{slot['slot_name']} — Lending Report — {monthLabel(month_year)}", TITLE_FONT, PatternFill("solid", fgColor="FFFFFF"), 28)
    r += 2

    # ── SECTION 1: LEDGER ──
    merge_title(r, "📋  CUSTOMER LEDGER", SECTION_FONT, SECTION_FILL)
    r += 1

    headers = ["Sl#", "Opening Date", "Customer Name", "Funding", f"Profit ({int(profit_rate*100)}%)",
               "Staff Comm (8%)", "Total Due", "Payment Done", "Balance", "Net Income", "Status", ""]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[r].height = 20
    r += 1

    ledger_start = r
    for c in customers:
        vals = [
            c["sl_no"], str(c["opening_date"]), c["customer_name"], float(c["funding"] or 0),
            float(c["profit_28_percent"]), float(c["staff_commission"]), float(c["total_payment_to_be_made"]),
            float(c["payment_has_been_done"] or 0), float(c["balance_recovery"]), float(c["net_income"]),
            "Closed" if c["is_closed"] else "Active", "",
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            if col in (4,5,6,7,8,9,10):
                cell.number_format = NUM_FMT
                cell.alignment = RIGHT
            elif col in (1,11,12):
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
            if c["is_closed"]:
                cell.fill = PatternFill("solid", fgColor="F1F2F4")
        r += 1
    ledger_end = r - 1

    for col, label in [(1, "TOTAL"), (4, f"=SUM(D{ledger_start}:D{ledger_end})"),
                        (5, f"=SUM(E{ledger_start}:E{ledger_end})"), (6, f"=SUM(F{ledger_start}:F{ledger_end})"),
                        (7, f"=SUM(G{ledger_start}:G{ledger_end})"), (8, f"=SUM(H{ledger_start}:H{ledger_end})"),
                        (9, f"=SUM(I{ledger_start}:I{ledger_end})"), (10, f"=SUM(J{ledger_start}:J{ledger_end})")]:
        cell = ws.cell(row=r, column=col, value=label)
        cell.font = BOLD_BODY
        cell.fill = TOTAL_FILL
        cell.border = BORDER
        if col != 1:
            cell.number_format = NUM_FMT
            cell.alignment = RIGHT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    for col in range(1, ncols + 1):
        ws.cell(row=r, column=col).fill = TOTAL_FILL
        ws.cell(row=r, column=col).border = BORDER
    r += 3

    # ── SECTION 2: DAILY COLLECTIONS ──
    merge_title(r, "📅  DAILY LEDGER — DEVANAM / PAWANA", SECTION_FONT, SECTION_FILL)
    r += 1

    d_headers = ["S.No", "Date", "Opening Pawana", "Income", "Khata", "Closing Pawana", "", "", "", "", "", ""]
    for c, h in enumerate(d_headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[r].height = 20
    r += 1

    daily_start = r
    for dr in daily_rows:
        date_str = f"{dr['day']:02d}.{int(m):02d}.{y[2:]}"
        vals = [dr["day"], date_str, dr["opening_pawana"], dr["income"], dr["khata"], dr["closing_pawana"]]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            if col in (3,4,5,6):
                cell.number_format = NUM_FMT
                cell.alignment = RIGHT
            else:
                cell.alignment = CENTER
        r += 1
    daily_end = r - 1

    for col, label in [(1, "TOTAL"), (4, f"=SUM(D{daily_start}:D{daily_end})"),
                        (5, f"=SUM(E{daily_start}:E{daily_end})")]:
        cell = ws.cell(row=r, column=col, value=label)
        cell.font = BOLD_BODY
        cell.fill = TOTAL_FILL
        cell.border = BORDER
        if col != 1:
            cell.number_format = NUM_FMT
            cell.alignment = RIGHT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    for col in range(1, ncols + 1):
        ws.cell(row=r, column=col).fill = TOTAL_FILL
        ws.cell(row=r, column=col).border = BORDER
    r += 3

    # ── SECTION 3: SUMMARY ──
    merge_title(r, "📊  MONTHLY SUMMARY", SECTION_FONT, SECTION_FILL)
    r += 1

    actual_recovery   = float(summary_row.get("actual_recovery") or 0)
    total_fund_profit = float(summary_row.get("total_funding_and_profit") or 0)
    live_balance      = float(summary_row.get("total_balance_recovery") or 0) - total_daily_income
    recovery_pct      = min((actual_recovery / total_fund_profit * 100), 100) if total_fund_profit else 0

    summary_pairs = [
        ("This Month Funding", float(summary_row.get("this_month_funding") or 0)),
        ("Total Running Funding", float(summary_row.get("running_funding") or 0)),
        ("Closed Funding", float(summary_row.get("closed_funding_total") or 0)),
        ("Total Funding + Profit", total_fund_profit),
        ("Actual Recovery (Upto Previous Month)", actual_recovery),
        ("Total Income Collected In This Month", total_daily_income),
        ("Live Balance (after daily)", live_balance),
        (f"Total Profit ({int(profit_rate*100)}%)", float(summary_row.get("total_profit_28") or 0)),
        ("Staff Commission (8%)", float(summary_row.get("total_staff_commission") or 0)),
        (f"Net Income ({int(net_rate*100)}%)", float(summary_row.get("total_net_income") or 0)),
        ("Kothi Capital", float(slot.get("kothi_amount") or 0)),
        ("Recovery Progress", recovery_pct / 100),
    ]
    for label, val in summary_pairs:
        lbl_cell = ws.cell(row=r, column=1, value=label)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        lbl_cell.font = BODY_FONT
        lbl_cell.alignment = LEFT
        val_cell = ws.cell(row=r, column=9, value=val)
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=12)
        val_cell.font = BOLD_BODY
        val_cell.alignment = RIGHT
        val_cell.number_format = PCT_FMT if label == "Recovery Progress" else NUM_FMT
        for col in range(1, ncols + 1):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT if (r % 2 == 0) else "FFFFFF")
        r += 1

    # ── Column widths ──
    widths = [8, 13, 22, 12, 13, 12, 12, 13, 12, 12, 10, 4]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_slot = "".join(ch if ch.isalnum() else "_" for ch in slot["slot_name"])
    filename = f"{safe_slot}_{month_year}.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ─── CUSTOMERS ────────────────────────────────────────────────────────────────
@app.route("/api/customers/<int:slot_id>/<month_year>", methods=["GET"])
@login_required
def get_customers(slot_id, month_year):
    profit_rate = get_profit_rate(slot_id)
    staff_rate  = 0.08
    net_rate    = profit_rate - staff_rate
    total_rate  = 1 + profit_rate

    rows = query(
        "SELECT * FROM customers WHERE slot_id=%s AND month_year=%s ORDER BY sl_no ASC",
        (slot_id, month_year),
    )
    for r in rows:
        if r.get("opening_date"):
            r["opening_date"] = str(r["opening_date"])
        f = float(r.get("funding") or 0)
        p = float(r.get("payment_has_been_done") or 0)
        r["profit_28_percent"]        = round(f * profit_rate, 2)
        r["staff_commission"]         = round(f * staff_rate, 2)
        r["net_income"]               = round(f * net_rate, 2)
        r["total_payment_to_be_made"] = round(f * total_rate, 2)
        r["balance_recovery"]         = round(f * total_rate - p, 2)
        r["profit_rate"]              = profit_rate

    return app.response_class(json.dumps(rows, default=decimal_default), mimetype="application/json")

@app.route("/api/customers", methods=["POST"])
@admin_required
def add_customer():
    d = request.json
    slot_id               = d.get("slot_id")
    customer_name         = d.get("customer_name")
    funding               = d.get("funding")
    opening_date          = d.get("opening_date")
    month_year            = d.get("month_year")
    payment_has_been_done = d.get("payment_has_been_done", 0) or 0
    daily_recovery        = d.get("daily_recovery", 0) or 0
    sl_no                 = d.get("sl_no")

    if not all([slot_id, customer_name, funding, opening_date, month_year]):
        return jsonify({"error": "Missing required fields"}), 400

    if not sl_no:
        row = query(
            "SELECT COALESCE(MAX(sl_no), 0) + 1 AS next_sl FROM customers WHERE slot_id=%s AND month_year=%s",
            (slot_id, month_year),
            fetchone=True,
        )
        sl_no = row["next_sl"]

    last_id = query(
        """INSERT INTO customers (slot_id, sl_no, opening_date, customer_name, funding,
           payment_has_been_done, month_year, daily_recovery)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
        (slot_id, sl_no, opening_date, customer_name, funding,
         payment_has_been_done, month_year, daily_recovery),
        commit=True,
    )
    new_row = query("SELECT * FROM customers WHERE id=%s", (last_id,), fetchone=True)
    if new_row:
        if new_row.get("opening_date"):
            new_row["opening_date"] = str(new_row["opening_date"])
        profit_rate = get_profit_rate(slot_id)
        staff_rate  = 0.08
        net_rate    = profit_rate - staff_rate
        total_rate  = 1 + profit_rate
        f = float(new_row.get("funding") or 0)
        p = float(new_row.get("payment_has_been_done") or 0)
        new_row["profit_28_percent"]        = round(f * profit_rate, 2)
        new_row["staff_commission"]         = round(f * staff_rate, 2)
        new_row["net_income"]               = round(f * net_rate, 2)
        new_row["total_payment_to_be_made"] = round(f * total_rate, 2)
        new_row["balance_recovery"]         = round(f * total_rate - p, 2)
        new_row["profit_rate"]              = profit_rate

    return app.response_class(
        json.dumps(new_row, default=decimal_default),
        status=201,
        mimetype="application/json"
    )

# ─── REORDER CUSTOMERS — must be BEFORE /api/customers/<int:cid> routes ──────
@app.route("/api/customers/reorder", methods=["POST"])
@admin_required
def reorder_customers():
    data       = request.json
    slot_id    = data.get("slot_id")
    month_year = data.get("month_year")
    order      = data.get("order", [])

    if not order:
        return jsonify({"error": "Missing order list"}), 400

    for index, cid in enumerate(order, start=1):
        query(
            "UPDATE customers SET sl_no=%s WHERE id=%s AND slot_id=%s AND month_year=%s",
            (index, cid, slot_id, month_year),
            commit=True,
        )
    return jsonify({"message": "Reordered successfully", "count": len(order)})

# ─── These routes use <int:cid> so they MUST come after /reorder ─────────────
@app.route("/api/customers/<int:cid>", methods=["PUT"])
@admin_required
def update_customer(cid):
    d = request.json or {}
    fields = []
    values = []
    allowed = ["payment_has_been_done", "daily_recovery", "is_closed",
               "customer_name", "funding", "opening_date"]
    numeric_fields = {"payment_has_been_done", "daily_recovery", "funding", "is_closed"}

    for key in allowed:
        if key not in d:
            continue
        val = d[key]

        if key in numeric_fields:
            if val is None or (isinstance(val, str) and val.strip() == ""):
                return jsonify({"error": f"{key} cannot be empty"}), 400
            try:
                val = float(val) if key != "is_closed" else int(val)
            except (TypeError, ValueError):
                return jsonify({"error": f"{key} must be a number"}), 400

        if key == "customer_name" and isinstance(val, str) and val.strip() == "":
            return jsonify({"error": "customer_name cannot be empty"}), 400

        fields.append(f"{key}=%s")
        values.append(val)

    if not fields:
        return jsonify({"error": "No fields to update"}), 400

    values.append(cid)
    try:
        query(f"UPDATE customers SET {', '.join(fields)} WHERE id=%s", values, commit=True)
    except mysql.connector.Error as e:
        return jsonify({"error": f"Database error: {e.msg}"}), 400

    updated = query("SELECT * FROM customers WHERE id=%s", (cid,), fetchone=True)
    if not updated:
        return jsonify({"error": "Customer not found"}), 404
    if updated.get("opening_date"):
        updated["opening_date"] = str(updated["opening_date"])

    slot_id     = updated.get("slot_id")
    profit_rate = get_profit_rate(slot_id)
    staff_rate  = 0.08
    net_rate    = profit_rate - staff_rate
    total_rate  = 1 + profit_rate
    f = float(updated.get("funding") or 0)
    p = float(updated.get("payment_has_been_done") or 0)
    updated["profit_28_percent"]        = round(f * profit_rate, 2)
    updated["staff_commission"]         = round(f * staff_rate, 2)
    updated["net_income"]               = round(f * net_rate, 2)
    updated["total_payment_to_be_made"] = round(f * total_rate, 2)
    updated["balance_recovery"]         = round(f * total_rate - p, 2)
    updated["profit_rate"]              = profit_rate

    return app.response_class(json.dumps(updated, default=decimal_default), mimetype="application/json")

@app.route("/api/customers/<int:cid>", methods=["DELETE"])
@admin_required
def delete_customer(cid):
    query("DELETE FROM customers WHERE id=%s", (cid,), commit=True)
    return jsonify({"message": "Deleted successfully"})


# ─── DAILY COLLECTIONS ───────────────────────────────────────────────────────
@app.route("/api/daily-collections/<int:slot_id>/<month_year>", methods=["GET"])
@login_required
def get_daily_collections(slot_id, month_year):
    collections = query(
        """SELECT *, DAY(collection_date) AS day_number, actual_collected AS amount_collected
           FROM daily_collections
           WHERE slot_id=%s AND month_year=%s
           ORDER BY collection_date ASC""",
        (slot_id, month_year),
    )

    opening_row = query(
        "SELECT opening_amount FROM opening_pawana WHERE slot_id=%s AND month_year=%s",
        (slot_id, month_year),
        fetchone=True,
    )
    opening_amount = float(opening_row["opening_amount"]) if opening_row else 0

    coll_map = {r["day_number"]: r for r in collections}

    y, m = month_year.split("-")
    days_in_month = calendar.monthrange(int(y), int(m))[1]

    rows = []
    running_pawana = opening_amount
    for day in range(1, days_in_month + 1):
        rec    = coll_map.get(day)
        income = float(rec["amount_collected"]) if rec else 0
        khata  = float(rec["khata_amount"]) if rec and rec.get("khata_amount") is not None else 0
        opening_pawana_for_day = running_pawana
        closing_pawana_for_day = opening_pawana_for_day + income - khata

        rows.append({
            "day_number":       day,
            "amount_collected": income,
            "khata_amount":     khata,
            "opening_pawana":   opening_pawana_for_day,
            "closing_pawana":   closing_pawana_for_day,
            "note":             rec.get("note") if rec else "",
        })
        running_pawana = closing_pawana_for_day

    return app.response_class(
        json.dumps({"opening_amount": opening_amount, "days": rows}, default=decimal_default),
        mimetype="application/json"
    )

@app.route("/api/daily-collections", methods=["POST"])
@admin_required
def upsert_daily_collection():
    d          = request.json
    slot_id    = d.get("slot_id")
    month_year = d.get("month_year")
    day_number = d.get("day_number")
    amount     = d.get("amount", 0) or 0
    khata      = d.get("khata_amount", None)
    note       = d.get("note", "") or ""
    y, m = month_year.split("-")
    collection_date = f"{y}-{m}-{str(day_number).zfill(2)}"

    if khata is None:
        existing = query(
            "SELECT khata_amount FROM daily_collections WHERE slot_id=%s AND collection_date=%s",
            (slot_id, collection_date), fetchone=True,
        )
        khata = existing["khata_amount"] if existing and existing.get("khata_amount") is not None else 0
    khata = khata or 0

    query(
        """INSERT INTO daily_collections (slot_id, month_year, collection_date, actual_collected, khata_amount, note)
           VALUES (%s, %s, %s, %s, %s, %s)
           ON DUPLICATE KEY UPDATE actual_collected=%s, khata_amount=%s, note=%s""",
        (slot_id, month_year, collection_date, amount, khata, note, amount, khata, note),
        commit=True,
    )
    row = query(
        """SELECT *, DAY(collection_date) AS day_number, actual_collected AS amount_collected
           FROM daily_collections
           WHERE slot_id=%s AND collection_date=%s""",
        (slot_id, collection_date), fetchone=True
    )
    return app.response_class(json.dumps(row, default=decimal_default), mimetype="application/json")

@app.route("/api/daily-collections/khata", methods=["POST"])
@admin_required
def upsert_khata():
    d          = request.json
    slot_id    = d.get("slot_id")
    month_year = d.get("month_year")
    day_number = d.get("day_number")
    khata      = d.get("khata_amount", 0) or 0
    y, m = month_year.split("-")
    collection_date = f"{y}-{m}-{str(day_number).zfill(2)}"
    query(
        """INSERT INTO daily_collections (slot_id, month_year, collection_date, actual_collected, khata_amount)
           VALUES (%s, %s, %s, 0, %s)
           ON DUPLICATE KEY UPDATE khata_amount=%s""",
        (slot_id, month_year, collection_date, khata, khata),
        commit=True,
    )
    row = query(
        """SELECT *, DAY(collection_date) AS day_number, actual_collected AS amount_collected
           FROM daily_collections
           WHERE slot_id=%s AND collection_date=%s""",
        (slot_id, collection_date), fetchone=True
    )
    return app.response_class(json.dumps(row, default=decimal_default), mimetype="application/json")

# ─── OPENING PAWANA ──────────────────────────────────────────────────────────
@app.route("/api/opening-pawana/<int:slot_id>/<month_year>", methods=["GET"])
@login_required
def get_opening_pawana(slot_id, month_year):
    row = query(
        "SELECT * FROM opening_pawana WHERE slot_id=%s AND month_year=%s",
        (slot_id, month_year),
        fetchone=True,
    )
    return app.response_class(
        json.dumps(row or {"opening_amount": 0}, default=decimal_default),
        mimetype="application/json"
    )

@app.route("/api/opening-pawana", methods=["POST"])
@admin_required
def save_opening_pawana():
    d              = request.json
    slot_id        = d.get("slot_id")
    month_year     = d.get("month_year")
    opening_amount = d.get("opening_amount", 0) or 0
    query(
        """INSERT INTO opening_pawana (slot_id, month_year, opening_amount)
           VALUES (%s, %s, %s)
           ON DUPLICATE KEY UPDATE opening_amount=%s""",
        (slot_id, month_year, opening_amount, opening_amount),
        commit=True,
    )
    row = query(
        "SELECT * FROM opening_pawana WHERE slot_id=%s AND month_year=%s",
        (slot_id, month_year),
        fetchone=True,
    )
    return app.response_class(json.dumps(row, default=decimal_default), mimetype="application/json")

# ─── CARRY FORWARD ───────────────────────────────────────────────────────────
@app.route("/api/carry-forward", methods=["POST"])
@admin_required
def carry_forward():
    data       = request.json
    slot_id    = data.get("slot_id")
    from_month = data.get("from_month")
    to_month   = data.get("to_month")

    if not all([slot_id, from_month, to_month]):
        return jsonify({"error": "Missing slot_id, from_month or to_month"}), 400

    active = query(
        "SELECT * FROM customers WHERE slot_id=%s AND month_year=%s AND is_closed=0",
        (slot_id, from_month),
    )

    if not active:
        return jsonify({"copied": 0, "message": "No active customers to carry forward."})

    copied  = 0
    skipped = 0
    for c in active:
        exists = query(
            "SELECT id FROM customers WHERE slot_id=%s AND month_year=%s AND customer_name=%s",
            (slot_id, to_month, c["customer_name"]),
            fetchone=True,
        )
        if exists:
            skipped += 1
            continue

        row = query(
            "SELECT COALESCE(MAX(sl_no), 0) + 1 AS next_sl FROM customers WHERE slot_id=%s AND month_year=%s",
            (slot_id, to_month),
            fetchone=True,
        )
        next_sl = row["next_sl"]

        query(
            """INSERT INTO customers
               (slot_id, sl_no, opening_date, customer_name, funding,
                payment_has_been_done, month_year, daily_recovery, is_closed)
               VALUES (%s, %s, %s, %s, %s, 0, %s, %s, 0)""",
            (slot_id, next_sl, c["opening_date"], c["customer_name"],
             c["funding"], to_month, c["daily_recovery"]),
            commit=True,
        )
        copied += 1

    return jsonify({
        "copied":  copied,
        "skipped": skipped,
        "message": f"{copied} customer(s) carried to {to_month}. {skipped} already existed and were skipped."
    })


if __name__ == "__main__":
    app.run(host='0.0.0.0', debug=True, port=8000)